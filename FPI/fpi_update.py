"""
NSDL FPI Fortnightly Sector-wise Data -- automated updater.

Run this script and it will:
  1. Read FPI_Sectorial-Analysis.xlsx in the same folder.
  2. Discover the list of fortnightly reports NSDL has published.
  3. Download any new ones that aren't yet in the workbook.
  4. Update the workbook (Auto Data + Sectors Wide + Monthly + Quarterly + Yearly + existing sheet).
  5. Refresh fpi_data.js next to the workbook so FPI_Dashboard.html sees fresh numbers.

Idempotent -- re-running is safe. On no new data it exits in a second.

Dependencies:  pip install requests beautifulsoup4 openpyxl
"""
from __future__ import annotations
import json, re, sys, time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Callable

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule

HERE = Path(__file__).resolve().parent
WORKBOOK_PATH = HERE / "FPI_Sectorial-Analysis.xlsx"
JSON_OUT_PATH = HERE / "fpi_data.json"
JS_OUT_PATH = HERE / "fpi_data.js"
LOG_PATH = HERE / "fpi_update.log"

BASE_URL = "https://www.fpi.nsdl.co.in/web"
SELECTION_URL = f"{BASE_URL}/Reports/FPI_Fortnightly_Selection.aspx"
REPORT_PREFIX = f"{BASE_URL}/StaticReports/Fortnightly_Sector_wise_FII_Investment_Data/"
USER_AGENT = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

EARLIEST_FORTNIGHT_END = date(2024, 7, 15)
MONTHS_MAP = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,"JUNE":6,
              "JUL":7,"JULY":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}

# ------- log -------
def log(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    # Ensure ASCII only -- avoid Windows cp1252 console / sync errors
    try:
        sys.stdout.write(line + "\n")
        sys.stdout.flush()
    except Exception:
        pass
    try:
        with LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass

# ------- dataclasses -------
@dataclass
class Fortnight:
    end_date: date
    period_start: date
    period_end: date
    label: str
    report_url: str

@dataclass
class SectorRow:
    sector: str
    equity_inr: Optional[float] = None
    debt_general_inr: Optional[float] = None
    debt_vrr_inr: Optional[float] = None
    debt_far_inr: Optional[float] = None
    hybrid_inr: Optional[float] = None
    mf_equity_inr: Optional[float] = None
    mf_debt_inr: Optional[float] = None
    mf_hybrid_inr: Optional[float] = None
    mf_solution_inr: Optional[float] = None
    mf_other_inr: Optional[float] = None
    aif_inr: Optional[float] = None
    total_inr: Optional[float] = None

# ------- http session -------
def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    try:
        s.get(SELECTION_URL, timeout=30)
    except Exception as e:
        log(f"WARN: failed to warm session: {e}")
    return s

# ------- discovery -------
_FILENAME_RE = re.compile(r"FIIInvestSector_([A-Za-z]+)(\d{1,2})(\d{4})\.html", re.IGNORECASE)

def parse_report_filename(name: str) -> Optional[date]:
    m = _FILENAME_RE.search(name)
    if not m: return None
    mon_str, day_str, year_str = m.groups()
    mon = MONTHS_MAP.get(mon_str.upper())
    if not mon: return None
    try: return date(int(year_str), mon, int(day_str))
    except ValueError: return None

def period_start_for(end: date) -> date:
    return end.replace(day=1) if end.day == 15 else end.replace(day=16)

def period_label(end: date) -> str:
    return f"{end.strftime('%b')} {period_start_for(end).day:02d}-{end.day:02d}, {end.year}"

def discover_fortnights(session: requests.Session) -> List[Fortnight]:
    r = session.get(SELECTION_URL, timeout=30); r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    dropdown = soup.find("select", {"id": "ddlfortnighly"}) or soup.find("select")
    if not dropdown: raise RuntimeError("Selection dropdown not found on NSDL page")
    out: List[Fortnight] = []
    for opt in dropdown.find_all("option"):
        rel = (opt.get("value") or "").strip()
        if not rel: continue
        fname = rel.split("/")[-1]
        end = parse_report_filename(fname)
        if not end or end < EARLIEST_FORTNIGHT_END: continue
        out.append(Fortnight(end, period_start_for(end), end, period_label(end), REPORT_PREFIX + fname))
    out.sort(key=lambda f: f.end_date)
    return out

# ------- parsing -------
def _to_num(text: str) -> Optional[float]:
    if text is None: return None
    t = text.strip().replace(",", "").replace("\xa0", "")
    if t in ("", "-"): return None
    try: return float(t)
    except ValueError: return None

CURR_NET_INV_BLOCK = 2
PREV_NET_INV_BLOCK = 1
LEAD_COLS = 2

def _schema_for(row_len: int) -> Tuple[int, int]:
    """Two NSDL schemas. Returns (inr_cols_per_block, block_total_cells)."""
    if row_len >= LEAD_COLS + 24 * 4:  # 98 (NEW: Aug 31, 2024+)
        return 12, 24
    if row_len >= LEAD_COLS + 10 * 4:  # 42 (OLD: pre Aug 31, 2024)
        return 5, 10
    return 0, 0

def _block_inr_slice(row_cells, block_idx, inr_cols, block_total):
    start = LEAD_COLS + block_idx * block_total
    return row_cells[start:start + inr_cols]

def parse_report(html: str) -> Dict[str, Dict[str, SectorRow]]:
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find_all("table")[0]
    rows = table.find_all("tr")
    result = {"prev": {}, "curr": {}}
    for tr in rows[4:]:
        cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
        inr_cols, block_total = _schema_for(len(cells))
        if inr_cols == 0: continue
        sector = cells[1]
        if not sector: continue
        # NSDL totals row is labeled "Total" or "Grand Total" depending on schema -- normalize
        sl = sector.lower().strip()
        if sl.startswith("total") or "grand total" in sl:
            sector_key = "TOTAL"
        else:
            sector_key = sector
        for key, block_idx in (("prev", PREV_NET_INV_BLOCK), ("curr", CURR_NET_INV_BLOCK)):
            v = _block_inr_slice(cells, block_idx, inr_cols, block_total)
            if inr_cols == 12:
                row = SectorRow(
                    sector=sector_key,
                    equity_inr=_to_num(v[0]), debt_general_inr=_to_num(v[1]),
                    debt_vrr_inr=_to_num(v[2]), debt_far_inr=_to_num(v[3]),
                    hybrid_inr=_to_num(v[4]), mf_equity_inr=_to_num(v[5]),
                    mf_debt_inr=_to_num(v[6]), mf_hybrid_inr=_to_num(v[7]),
                    mf_solution_inr=_to_num(v[8]), mf_other_inr=_to_num(v[9]),
                    aif_inr=_to_num(v[10]), total_inr=_to_num(v[11]),
                )
            else:
                row = SectorRow(
                    sector=sector_key,
                    equity_inr=_to_num(v[0]), debt_general_inr=_to_num(v[1]),
                    debt_vrr_inr=_to_num(v[2]), hybrid_inr=_to_num(v[3]),
                    total_inr=_to_num(v[4]),
                )
            result[key][sector_key] = row
    return result

def fetch_and_parse(session: requests.Session, fn: Fortnight) -> Dict[str, Dict[str, SectorRow]]:
    log(f"  fetching {fn.label} ... {fn.report_url}")
    r = session.get(fn.report_url, timeout=60); r.raise_for_status()
    return parse_report(r.text)

# ------- workbook state -------
AUTO_DATA_SHEET = "Auto Data"
WIDE_SHEET = "Sectors Wide"
MONTHLY_SHEET = "Monthly"
QUARTERLY_SHEET = "Quarterly"
YEARLY_SHEET = "Yearly (FY)"
EXISTING_USER_SHEET = "FPIs Fortnightly Investments"

AUTO_DATA_HEADERS = [
    "Fortnight End", "Period Label", "Sector",
    "Equity (INR Cr)", "Debt-General (INR Cr)", "Debt-VRR (INR Cr)", "Debt-FAR (INR Cr)",
    "Hybrid (INR Cr)", "MF-Equity (INR Cr)", "MF-Debt (INR Cr)", "MF-Hybrid (INR Cr)",
    "MF-Solution (INR Cr)", "MF-Other (INR Cr)", "AIF (INR Cr)", "Total (INR Cr)",
]

def ensure_auto_data_sheet(wb: Workbook) -> None:
    if AUTO_DATA_SHEET in wb.sheetnames: return
    ws = wb.create_sheet(AUTO_DATA_SHEET)
    ws.append(AUTO_DATA_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "D2"

def existing_fortnight_ends(wb: Workbook) -> set:
    if AUTO_DATA_SHEET not in wb.sheetnames: return set()
    ws = wb[AUTO_DATA_SHEET]
    ends = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if isinstance(v, datetime): ends.add(v.date())
        elif isinstance(v, date): ends.add(v)
        elif isinstance(v, str):
            try: ends.add(datetime.strptime(v[:10], "%Y-%m-%d").date())
            except Exception: pass
    return ends

def append_auto_data(wb: Workbook, fn: Fortnight, parsed_curr: Dict[str, SectorRow]) -> None:
    ws = wb[AUTO_DATA_SHEET]
    for sector, row in sorted(parsed_curr.items()):
        if sector == "TOTAL":
            continue  # don't store totals row
        ws.append([
            fn.end_date, fn.label, sector,
            row.equity_inr, row.debt_general_inr, row.debt_vrr_inr, row.debt_far_inr,
            row.hybrid_inr, row.mf_equity_inr, row.mf_debt_inr, row.mf_hybrid_inr,
            row.mf_solution_inr, row.mf_other_inr, row.aif_inr, row.total_inr,
        ])

# ------- aggregations -------
SECTOR_ORDER = [
    "Automobile and Auto Components","Capital Goods","Chemicals","Construction",
    "Construction Materials","Consumer Durables","Consumer Services",
    "Diversified","Fast Moving Consumer Goods","Financial Services","Forest Materials",
    "Healthcare","Information Technology","Media Entertainment & Publication",
    "Media, Entertainment & Publication","Metals & Mining",
    "Oil, Gas & Consumable Fuels","Power","Realty","Services","Sovereign","Telecommunication",
    "Textiles","Utilities","Others",
]

def _read_auto_data(wb: Workbook) -> List[dict]:
    if AUTO_DATA_SHEET not in wb.sheetnames: return []
    ws = wb[AUTO_DATA_SHEET]
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        end = rec.get("Fortnight End")
        if isinstance(end, datetime): rec["Fortnight End"] = end.date()
        out.append(rec)
    return out

def _fy_label(d: date) -> str:
    if d.month >= 4: return f"FY {d.year % 100:02d}-{(d.year + 1) % 100:02d}"
    return f"FY {(d.year - 1) % 100:02d}-{d.year % 100:02d}"
def _fy_sort_key(s: str): return int(s.split()[1].split("-")[0])
def _quarter_label(d: date) -> str:
    q = (d.month - 1) // 3 + 1; return f"{d.year} Q{q}"
def _quarter_sort_key(s: str):
    yr, q = s.split(" Q"); return (int(yr), int(q))
def _month_label(d: date) -> str: return d.strftime("%b %Y")
def _month_sort_key(s: str): return datetime.strptime(s, "%b %Y")

def rebuild_aggregations(wb: Workbook) -> None:
    records = _read_auto_data(wb)
    if not records: return
    # Strip any TOTAL rows from records used for pivots (we no longer store them, but be defensive)
    records = [r for r in records if (r.get("Sector") or "").strip().upper() != "TOTAL"
               and "grand total" not in (r.get("Sector") or "").lower()]
    fortnights = sorted({r["Fortnight End"] for r in records})
    sectors = sorted({r["Sector"] for r in records})
    ordered_sectors = [s for s in SECTOR_ORDER if s in sectors] + \
                      [s for s in sectors if s not in SECTOR_ORDER]
    wide: Dict[Tuple[str, date], Optional[float]] = {}
    for r in records:
        wide[(r["Sector"], r["Fortnight End"])] = r.get("Equity (INR Cr)")

    def _write_pivot(sheet_name, period_of, period_order):
        if sheet_name in wb.sheetnames: del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        period_to_fortnights: Dict[str, List[date]] = {}
        for fn in fortnights:
            p = period_of(fn)
            period_to_fortnights.setdefault(p, []).append(fn)
        periods = sorted(period_to_fortnights.keys(), key=period_order)
        c = ws.cell(1, 1, "Sector")
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F4E78")
        for j, p in enumerate(periods, start=2):
            c = ws.cell(1, j, p)
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center")
        for i, sec in enumerate(ordered_sectors, start=2):
            ws.cell(i, 1, sec)
            for j, p in enumerate(periods, start=2):
                vals = [wide.get((sec, fn)) for fn in period_to_fortnights[p]]
                vals = [v for v in vals if v is not None]
                ws.cell(i, j, sum(vals) if vals else None)
        total_row = len(ordered_sectors) + 2
        ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
        for j in range(2, len(periods) + 2):
            col = get_column_letter(j)
            ws.cell(total_row, j, f"=SUM({col}2:{col}{total_row - 1})").font = Font(bold=True)
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 36
        for j in range(2, len(periods) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 14
        last_col = get_column_letter(len(periods) + 1)
        rng = f"B2:{last_col}{total_row - 1}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=-8000, start_color="C00000",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=8000, end_color="2E7D32",
        ))

    _write_pivot(WIDE_SHEET, lambda d: d.isoformat(), lambda s: s)
    _write_pivot(MONTHLY_SHEET, _month_label, _month_sort_key)
    _write_pivot(QUARTERLY_SHEET, _quarter_label, _quarter_sort_key)
    _write_pivot(YEARLY_SHEET, _fy_label, _fy_sort_key)
    _add_trend_chart(wb)
    _add_monthly_bar_chart(wb)

def _add_trend_chart(wb: Workbook) -> None:
    if WIDE_SHEET not in wb.sheetnames: return
    ws = wb[WIDE_SHEET]
    max_col = ws.max_column; max_row = ws.max_row
    if max_col < 2 or max_row < 3: return
    chart = LineChart()
    chart.title = "Equity Net Investment by Sector (INR Cr) - top 8 sectors"
    chart.y_axis.title = "INR Cr"; chart.x_axis.title = "Fortnight"
    chart.width = 30; chart.height = 15
    sector_totals = []
    for i in range(2, max_row):
        name = ws.cell(i, 1).value
        vals = [ws.cell(i, j).value or 0 for j in range(2, max_col + 1)]
        sector_totals.append((name, sum(abs(v) for v in vals), i))
    top = sorted(sector_totals, key=lambda x: -x[1])[:8]
    for name, _, i in top:
        data_ref = Reference(ws, min_col=2, max_col=max_col, min_row=i, max_row=i)
        chart.add_data(data_ref, titles_from_data=False)
        chart.series[-1].tx = SeriesLabel(strRef=StrRef(f"'{WIDE_SHEET}'!$A${i}"))
    cats = Reference(ws, min_col=2, max_col=max_col, min_row=1, max_row=1)
    chart.set_categories(cats); chart.legend.position = "b"
    ws.add_chart(chart, f"{get_column_letter(max_col + 2)}2")

def _add_monthly_bar_chart(wb: Workbook) -> None:
    if MONTHLY_SHEET not in wb.sheetnames: return
    ws = wb[MONTHLY_SHEET]
    max_col = ws.max_column; max_row = ws.max_row
    if max_col < 2 or max_row < 3: return
    chart = BarChart(); chart.type = "col"; chart.style = 11
    chart.title = "Monthly Total FPI Equity Net Investment (INR Cr)"
    chart.y_axis.title = "INR Cr"; chart.x_axis.title = "Month"
    chart.width = 28; chart.height = 12
    total_row = max_row
    data_ref = Reference(ws, min_col=2, max_col=max_col, min_row=total_row, max_row=total_row)
    chart.add_data(data_ref, titles_from_data=False)
    cats = Reference(ws, min_col=2, max_col=max_col, min_row=1, max_row=1)
    chart.set_categories(cats); chart.legend = None
    ws.add_chart(chart, f"{get_column_letter(max_col + 2)}2")

# ------- existing user sheet -------
def update_existing_user_sheet(wb: Workbook, fn: Fortnight, parsed_curr: Dict[str, SectorRow]) -> None:
    if EXISTING_USER_SHEET not in wb.sheetnames: return
    ws = wb[EXISTING_USER_SHEET]
    header_row = 1
    existing_labels = set()
    for c in ws[header_row]:
        if c.value is None: continue
        existing_labels.add(str(c.value).replace("\n", " ").strip().lower())
    target = f"net investment {fn.label}".lower()
    if target in existing_labels: return
    insert_col = None
    for j in range(ws.max_column, 1, -1):
        v = ws.cell(header_row, j).value
        if v and "Net Investment" in str(v):
            insert_col = j + 1; break
    if insert_col is None: insert_col = ws.max_column + 1
    ws.insert_cols(insert_col)
    ws.cell(header_row, insert_col, f"Net Investment\n{fn.label}")
    ws.cell(header_row, insert_col).font = Font(bold=True)
    for r in range(header_row + 1, ws.max_row + 1):
        sector = ws.cell(r, 1).value
        if not sector: continue
        row = parsed_curr.get(str(sector).strip())
        if row and row.equity_inr is not None:
            ws.cell(r, insert_col, row.equity_inr)

# ------- exports -------
def export_json(wb: Workbook) -> None:
    records = _read_auto_data(wb)
    # filter out any TOTAL/Grand Total rows defensively
    records = [r for r in records if (r.get("Sector") or "").strip().upper() != "TOTAL"
               and "grand total" not in (r.get("Sector") or "").lower()]
    payload = []
    for r in records:
        end = r["Fortnight End"]
        if isinstance(end, datetime): end = end.date()
        payload.append({
            "end": end.isoformat() if end else None,
            "label": r.get("Period Label"), "sector": r.get("Sector"),
            "equity": r.get("Equity (INR Cr)"),
            "debt_general": r.get("Debt-General (INR Cr)"),
            "debt_vrr": r.get("Debt-VRR (INR Cr)"),
            "debt_far": r.get("Debt-FAR (INR Cr)"),
            "hybrid": r.get("Hybrid (INR Cr)"),
            "mf_equity": r.get("MF-Equity (INR Cr)"),
            "mf_debt": r.get("MF-Debt (INR Cr)"),
            "mf_hybrid": r.get("MF-Hybrid (INR Cr)"),
            "mf_solution": r.get("MF-Solution (INR Cr)"),
            "mf_other": r.get("MF-Other (INR Cr)"),
            "aif": r.get("AIF (INR Cr)"),
            "total": r.get("Total (INR Cr)"),
        })
    bundle = {"generated_at": datetime.now().isoformat(timespec="seconds"), "rows": payload}
    with JSON_OUT_PATH.open("w", encoding="utf-8") as f:
        json.dump(bundle, f, ensure_ascii=False)
    with JS_OUT_PATH.open("w", encoding="utf-8") as f:
        f.write("window.fpiData = "); json.dump(bundle, f, ensure_ascii=False); f.write(";\n")
    log(f"  wrote {JSON_OUT_PATH.name} & {JS_OUT_PATH.name} ({len(payload)} rows)")

# ------- main -------
def main() -> int:
    if not WORKBOOK_PATH.exists():
        log(f"ERROR: workbook not found: {WORKBOOK_PATH}"); return 1
    log(f"Loading workbook {WORKBOOK_PATH.name}")
    wb = load_workbook(WORKBOOK_PATH)
    ensure_auto_data_sheet(wb)
    session = make_session()
    log("Discovering fortnights ...")
    all_fortnights = discover_fortnights(session)
    log(f"  NSDL has {len(all_fortnights)} fortnights from {all_fortnights[0].label} to {all_fortnights[-1].label}")
    have = existing_fortnight_ends(wb)
    log(f"  workbook already has {len(have)} fortnights")
    new_ones = [fn for fn in all_fortnights if fn.end_date not in have]
    log(f"  {len(new_ones)} new fortnight(s) to fetch")
    fetched = 0
    for fn in new_ones:
        try:
            parsed = fetch_and_parse(session, fn)
            append_auto_data(wb, fn, parsed["curr"])
            update_existing_user_sheet(wb, fn, parsed["curr"])
            fetched += 1
            time.sleep(0.5)
        except Exception as e:
            log(f"  ERROR on {fn.label}: {e}")
    if fetched or (AUTO_DATA_SHEET in wb.sheetnames and wb[AUTO_DATA_SHEET].max_row > 1):
        log("Rebuilding aggregation sheets...")
        rebuild_aggregations(wb)
    wb.save(WORKBOOK_PATH)
    log(f"Saved workbook ({WORKBOOK_PATH})")
    export_json(wb)
    log(f"Done. New fortnights added this run: {fetched}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
